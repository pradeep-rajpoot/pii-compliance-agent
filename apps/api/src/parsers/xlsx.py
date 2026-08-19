from __future__ import annotations

from pathlib import Path

import openpyxl
import xlrd

from src.models.block import TextBlock
from src.models.enums import FileType
from src.models.locator import XlsxLocator
from src.parsers.common import ParsedDocument


def _parse_xlsx(path: Path) -> ParsedDocument:
    # Load once with data_only=True purely to read cached *computed* values
    # for formula cells (what a human actually sees) -- this is the text we
    # send to the LLM. We discard this workbook after building blocks.
    values_wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)

    # Load a second time in normal mode so formulas are preserved verbatim.
    # This workbook is handed to the correction agent (via meta) so cell
    # overwrites during masking don't destroy formulas in untouched cells.
    formula_wb = openpyxl.load_workbook(str(path), data_only=False)

    blocks: list[TextBlock] = []
    for sheet_name in values_wb.sheetnames:
        ws = values_wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if text == "":
                    continue
                blocks.append(
                    TextBlock(
                        id=f"{sheet_name}!{cell.coordinate}",
                        text=text,
                        locator=XlsxLocator(sheet=sheet_name, cell=cell.coordinate),
                    )
                )

    values_wb.close()

    return ParsedDocument(
        file_type=FileType.XLSX,
        blocks=blocks,
        meta={"workbook": formula_wb},
    )


def _parse_xls(path: Path) -> ParsedDocument:
    # Legacy binary format: xlrd is read-only, no writer counterpart is in
    # our dependency set, so there is no "formula-preserving workbook" to
    # stash for correction -- correction of .xls is a documented limitation
    # (see correction_agent.py / README).
    book = xlrd.open_workbook(str(path))

    blocks: list[TextBlock] = []
    for sheet in book.sheets():
        for row_idx in range(sheet.nrows):
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                if value in (None, ""):
                    continue
                text = str(value)
                coord = _xls_coord(row_idx, col_idx)
                blocks.append(
                    TextBlock(
                        id=f"{sheet.name}!{coord}",
                        text=text,
                        locator=XlsxLocator(sheet=sheet.name, cell=coord),
                    )
                )

    return ParsedDocument(
        file_type=FileType.XLS,
        blocks=blocks,
        meta={"workbook": None},
    )


def _xls_coord(row_idx: int, col_idx: int) -> str:
    """Convert 0-based (row, col) to an A1-style coordinate, matching
    openpyxl's Cell.coordinate convention, so downstream code (correction)
    can treat xls/xlsx locators uniformly."""

    col = col_idx
    letters = ""
    while True:
        col, rem = divmod(col, 26)
        letters = chr(65 + rem) + letters
        if col == 0:
            break
        col -= 1
    return f"{letters}{row_idx + 1}"


def parse(path: Path) -> ParsedDocument:
    if path.suffix.lower() == ".xls":
        return _parse_xls(path)
    return _parse_xlsx(path)
